import inspect
import logging
import softest
import csv
from openpyxl import Workbook, load_workbook

class Utils(softest.TestCase):
    def asserListItemText(self, list, value):
        for item in list:
            print("The Text is : " + item.text)
            self.soft_assert(self.assertEqual, item.text, value)
            # if item.text == value:
            if item.text !=None and item.text in value:
                print("Assert PASS")
            else:
                print("Assert FAILED")
        self.assert_all()

    def custom_logger(logLevel=logging.DEBUG):
        # Set Class/Method Name from where it will call
        logger_name = inspect.stack()[1][3]
        # Create Logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        # Create console handler
        fl = logging.FileHandler("automation.log", mode='a')
        # Create Formatter how you want to logs to be formatted
        formatter = logging.Formatter('%(asctime)s - %(levelname)s : %(message)s', datefmt='%d/%m/%Y %I:%M:%S %p')
        # Add formatter to console or file handler
        fl.setFormatter(formatter)
        # Add console handler to logger
        logger.addHandler(fl)
        return logger

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        rowcount = sh.max_row
        columncount = sh.max_column

        for i in range(2, rowcount + 1):
            rowlist=[]
            for j in range(1, columncount + 1):
                rowlist.append(sh.cell(row=i, column=j).value)
            datalist.append(rowlist)
        return datalist

    def read_data_from_csv(file_name):
        datalist=[]
        csvdata = open(file_name, "r")
        reader = csv.reader(csvdata)
        next(reader)
        for rows in reader:
            datalist.append(rows)
        return datalist
